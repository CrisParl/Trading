import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, numbers

# Download ES futures data
ticker = 'ES=F'
start_date = '2025-01-01'
end_date = datetime.today().strftime('%Y-%m-%d')
data = yf.download(ticker, start=start_date, end=end_date, auto_adjust=False)

# Flatten MultiIndex columns if needed
if isinstance(data.columns, pd.MultiIndex):
    data.columns = [col[0] for col in data.columns]

# Required columns check
required = ['Open', 'High', 'Low', 'Close']
if all(col in data.columns for col in required):
    data = data.dropna(subset=required)
    data['ES_Range'] = data['High'] - data['Low']
    data['Open_Close_Diff'] = data['Close'] - data['Open']
    data['Direction'] = data['Open_Close_Diff'].apply(lambda x: '▲' if x > 0 else '▼')

    data = data[['Open', 'Close', 'High', 'Low', 'ES_Range', 'Open_Close_Diff', 'Direction']].copy()
    data.columns = ['ES_Open', 'ES_Close', 'ES_High', 'ES_Low', 'ES_Range', 'Open_Close_Diff', 'Direction']
    data.index.name = 'Date'
    data.reset_index(inplace=True)

    # Tag key news events
    news_map = {
        '2025-04-02': "Trump announces new tariffs",
        '2025-04-03': "China retaliates — crash continues",
        '2025-04-04': "Continued panic selling",
        '2025-04-07': "Trump denies tariff pause",
        '2025-04-09': "Tariff cut announced — market rallies",
        '2025-05-09': "US-China trade talks begin"
    }
    data['News_Event'] = data['Date'].dt.strftime('%Y-%m-%d').map(news_map).fillna('')

    # Add week/year columns
    data['Week'] = data['Date'].dt.isocalendar().week
    data['Year'] = data['Date'].dt.isocalendar().year

    # Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ES Data"

    headers = ['Date', 'ES_Open', 'ES_Close', 'ES_High', 'ES_Low', 'ES_Range', 'Open_Close_Diff', 'News_Event']
    ws.append(headers)

    # Fill styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    high_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    low_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    bold_font = Font(bold=True)

    # Write rows grouped by week
    for (year, week), group in data.groupby(['Year', 'Week']):
        first = group.iloc[0]
        last = group.iloc[-1]

        for _, row in group.iterrows():
            diff = f"{row['Direction']} {abs(row['Open_Close_Diff']):.2f}"
            ws.append([
                row['Date'], row['ES_Open'], row['ES_Close'],
                row['ES_High'], row['ES_Low'], row['ES_Range'],
                diff, row['News_Event']
            ])

        # Weekly summary row
        net_change = last['ES_Close'] - first['ES_Open']
        net_dir = '▲' if net_change > 0 else '▼'
        net_text = f"{net_dir} {abs(net_change):.2f}"
        ws.append([
            "Week Summary", first['ES_Open'], last['ES_Close'],
            '', '', '', net_text, ''
        ])
        ws.append([])  # spacer

    # Apply formatting
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        if all(cell.value is None for cell in row): continue
        if isinstance(row[0].value, datetime):
            row[0].number_format = numbers.FORMAT_DATE_YYYYMMDD2
        if isinstance(row[3].value, (int, float)):
            row[3].fill = high_fill
        if isinstance(row[4].value, (int, float)):
            row[4].fill = low_fill
        if isinstance(row[5].value, (int, float)):
            row[5].fill = gray_fill
        if isinstance(row[6].value, str):
            if row[6].value.startswith("▲"):
                row[6].fill = green_fill
            elif row[6].value.startswith("▼"):
                row[6].fill = red_fill
        if isinstance(row[7].value, str) and row[7].value != '':
            row[7].font = bold_font  # bold news

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Save file
    filename = "ES_Range_Data_With_News.xlsx"
    wb.save(filename)
    print(f"✅ Excel saved: {filename}")

else:
    print("❌ Missing columns. Got:", list(data.columns))
